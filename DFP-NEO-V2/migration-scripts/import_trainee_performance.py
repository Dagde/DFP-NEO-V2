#!/usr/bin/env python3
"""
TraineePerformance Table Import Script
=======================================
Imports PT-051 assessment data from all 6 course spreadsheets into the
TraineePerformance database table via direct PostgreSQL insertion.

Courses handled:
  - ADF301, ADF302, ADF303, ADF306  (25-27 trainees, 74 events each)
  - FIC210, FIC211                   (4-19 trainees, 11 events each)

Key differences handled between ADF and FIC courses:
  - ADF:  date as Excel datetime object  | FIC: date as "DD-MM-YYYY" string
  - ADF:  duration as float (1.0, 1.5)   | FIC: duration as "1.0hr" string
  - ADF:  QFI col = "Yes" (boolean flag) | FIC: QFI col = instructor name
  - ADF:  Col 4 = trainee name           | FIC: Col 4 header = "traineeFullName"

Total expected: ~7,801 records across 125 trainees
"""

import os
import sys
import json
import time
import random
import string
import re
import hashlib
from datetime import datetime, date
import openpyxl
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# All spreadsheets to import
SPREADSHEETS = [
    ("/workspace/ADF301_All_Trainees_v2.xlsx", "ADF301"),
    ("/workspace/ADF302_All_Trainees_v2.xlsx", "ADF302"),
    ("/workspace/ADF303_All_Trainees.xlsx",    "ADF303"),
    ("/workspace/ADF306_All_Trainees.xlsx",    "ADF306"),
    ("/workspace/FIC210_All_Trainees_v2.xlsx", "FIC210"),
    ("/workspace/FIC211_All_Trainees.xlsx",    "FIC211"),
]

# Course type determines parsing strategy
ADF_COURSES = {"ADF301", "ADF302", "ADF303", "ADF306"}
FIC_COURSES  = {"FIC210", "FIC211"}

# Header row index (1-based in openpyxl)
HEADER_ROW = 2
# First data row
DATA_START_ROW = 3

# Column indices (1-based)
COL_EVENT_CODE       = 2   # B: "BGF MB1", "FIC GND1"
COL_EVENT_DESC       = 3   # C: Human-readable description
COL_TRAINEE_NAME     = 4   # D: Trainee full name
COL_EVENT_ID         = 5   # E: eventId (empty - auto-generated)
COL_FLIGHT_NUMBER    = 6   # F: flightNumber (empty - derived from Col B)
COL_DATE             = 7   # G: date
COL_INSTRUCTOR       = 8   # H: Instructor name
COL_DCO_RESULT       = 9   # I: DCO/DPCO/DNCO
COL_OVERALL_GRADE    = 10  # J: Overall grade (int)
COL_OVERALL_RESULT   = 11  # K: Pass/Fail
COL_START_TIME       = 12  # L: "HH:MM"
COL_DURATION         = 13  # M: Hours (ADF: float, FIC: "1.0hr")
COL_END_TIME         = 14  # N: "HH:MM"
COL_IS_COMPLETED     = 15  # O: Yes/No
COL_QFI              = 16  # P: QFI comment (ADF: "Yes", FIC: instructor name)
COL_WEATHER          = 17  # Q: Weather description
COL_PROFILE          = 18  # R: Profile description
COL_OVERALL_COMMENT  = 19  # S: Overall comment (contains duplicate grade number - treat as null)
COL_NEST             = 20  # T: NEST comment (always 0 - no NEST times - treat as null)

# Ground school assessment columns (after all 22 elements)
COL_IS_ASSESSMENT    = 109  # DE: isAssessment (Yes/No)
COL_GS_RESULT        = 110  # DF: Ground school result percentage

# Element structure: each element has 4 columns:
# [category_col, element_col, score_col, comment_col]
# Score columns: 23, 27, 31, 35, 39, 43, 47, 51, 55, 59, 63, 67, 71, 75, 79, 83, 87, 91, 95, 99, 103, 107
ELEMENTS = [
    {"name": "Airmanship",             "category": "Core Dimensions",       "score_col": 23,  "comment_col": 24},
    {"name": "Preparation",            "category": "Core Dimensions",       "score_col": 27,  "comment_col": 28},
    {"name": "Technique",              "category": "Core Dimensions",       "score_col": 31,  "comment_col": 32},
    {"name": "Pre-Post Flight",        "category": "Procedural Framework",  "score_col": 35,  "comment_col": 36},
    {"name": "Walk Around",            "category": "Procedural Framework",  "score_col": 39,  "comment_col": 40},
    {"name": "Strap-in",               "category": "Procedural Framework",  "score_col": 43,  "comment_col": 44},
    {"name": "Ground Checks",          "category": "Procedural Framework",  "score_col": 47,  "comment_col": 48},
    {"name": "Airborne Checks",        "category": "Procedural Framework",  "score_col": 51,  "comment_col": 52},
    {"name": "Stationary",             "category": "Takeoff",               "score_col": 55,  "comment_col": 56},
    {"name": "Visual",                 "category": "Departure",             "score_col": 59,  "comment_col": 60},
    {"name": "Effects of Control",     "category": "Core Handling Skills",  "score_col": 63,  "comment_col": 64},
    {"name": "Trimming",               "category": "Core Handling Skills",  "score_col": 67,  "comment_col": 68},
    {"name": "Straight and Level",     "category": "Core Handling Skills",  "score_col": 71,  "comment_col": 72},
    {"name": "Level Medium Turn",      "category": "Turns",                 "score_col": 75,  "comment_col": 76},
    {"name": "Level Steep Turn",       "category": "Turns",                 "score_col": 79,  "comment_col": 80},
    {"name": "Visual - Initial & Pitch","category": "Recovery",             "score_col": 83,  "comment_col": 84},
    {"name": "Landing",                "category": "Landing",               "score_col": 87,  "comment_col": 88},
    {"name": "Crosswind",              "category": "Landing",               "score_col": 91,  "comment_col": 92},
    {"name": "Radio Comms",            "category": "Domestics",             "score_col": 95,  "comment_col": 96},
    {"name": "Situational Awareness",  "category": "Domestics",             "score_col": 99,  "comment_col": 100},
    {"name": "Lookout",                "category": "Domestics",             "score_col": 103, "comment_col": 104},
    {"name": "Knowledge",              "category": "Domestics",             "score_col": 107, "comment_col": 108},
]

# Valid grade values (as strings - matching Pt051Grade TypeScript type)
VALID_ELEMENT_GRADES  = {"MIN", "DEMO", "0", "1", "2", "3", "4", "5"}
VALID_OVERALL_GRADES  = {"No Grade", "0", "1", "2", "3", "4", "5"}
VALID_DCO_RESULTS     = {"DCO", "DPCO", "DNCO", ""}
VALID_OVERALL_RESULTS = {"P", "F", None}

# ---------------------------------------------------------------------------
# Utility Functions
# ---------------------------------------------------------------------------

def generate_event_id() -> str:
    """
    Generate a unique event ID in the same format as the app's scheduled events.
    Format: sched_[13-digit-timestamp][random-8-char-alphanumeric]
    """
    ts = int(time.time() * 1000)  # millisecond timestamp
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=8))
    return f"sched_{ts}{rand}"

def generate_cuid() -> str:
    """Generate a cuid-compatible unique ID for the record primary key."""
    ts = int(time.time() * 1000)
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=20))
    return f"c{ts}{rand}"

def parse_time_to_decimal(value) -> float | None:
    """
    Convert HH:MM string to decimal hours.
    "08:00" -> 8.0,  "09:30" -> 9.5,  "14:15" -> 14.25
    Returns None for invalid/missing values.
    """
    if value is None:
        return None
    s = str(value).strip()
    if ':' in s:
        parts = s.split(':')
        try:
            hours   = int(parts[0])
            minutes = int(parts[1])
            return round(hours + minutes / 60.0, 4)
        except (ValueError, IndexError):
            return None
    try:
        return float(s)
    except ValueError:
        return None

def parse_duration(value, course: str) -> float | None:
    """
    Parse duration field.
    ADF courses: float or int (1, 1.2, 1.5)
    FIC courses: string like "1.0hr", "1.5hr"
    """
    if value is None:
        return None
    s = str(value).strip()
    # Remove "hr" suffix used in FIC courses
    s = re.sub(r'hr$', '', s, flags=re.IGNORECASE).strip()
    try:
        return float(s)
    except ValueError:
        return None

def parse_date(value, course: str) -> str | None:
    """
    Parse date field to ISO format YYYY-MM-DD.
    ADF courses: Excel datetime object (datetime.datetime)
    FIC courses: string "DD-MM-YYYY"
    Returns None for invalid/missing values.
    """
    if value is None:
        return None
    # ADF: datetime object from Excel
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    # FIC: "DD-MM-YYYY" format
    if re.match(r'^\d{2}-\d{2}-\d{4}$', s):
        try:
            return datetime.strptime(s, "%d-%m-%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    # Try ISO already
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return None

def parse_grade(value) -> str:
    """
    Convert grade value to string matching Pt051OverallGrade type.
    Handles: int 3 -> "3", float 3.0 -> "3", str "3" -> "3"
    Returns "No Grade" for null/invalid values.
    """
    if value is None:
        return "No Grade"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    if s in VALID_OVERALL_GRADES:
        return s
    # Try converting float string "3.0" -> "3"
    try:
        f = float(s)
        s = str(int(f))
        if s in VALID_OVERALL_GRADES:
            return s
    except ValueError:
        pass
    return "No Grade"

def parse_element_grade(value) -> str | None:
    """
    Convert element grade value to string matching Pt051Grade type.
    Handles: int 3 -> "3", float 3.0 -> "3", str "DEMO" -> "DEMO"
    Returns None for null/missing values.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip().upper()
    if s in {'MIN', 'DEMO'}:
        return s
    # Numeric grade
    try:
        n = int(float(s))
        ns = str(n)
        if ns in VALID_ELEMENT_GRADES:
            return ns
    except ValueError:
        pass
    return None

def parse_overall_result(value) -> str | None:
    """
    Convert overall result to 'P' or 'F'.
    "Pass" -> "P", "Fail" -> "F", None -> None
    """
    if value is None:
        return None
    s = str(value).strip().lower()
    if s in ('pass', 'p'):
        return 'P'
    if s in ('fail', 'f'):
        return 'F'
    return None

def parse_bool(value) -> bool:
    """Convert Yes/No/True/False/0/1 to boolean."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    return s in ('yes', 'true', '1')

def parse_ground_school_result(value) -> int | None:
    """Parse ground school percentage. Returns None for NaN/missing."""
    if value is None:
        return None
    try:
        f = float(value)
        if f != f:  # NaN check
            return None
        return int(f)
    except (ValueError, TypeError):
        return None

def build_comments(qfi, weather, profile, overall, nest, course: str) -> str | None:
    """
    Build the structured comment string expected by the app's parseComments() function.
    Format: "QFI: [text]\nWeather: [text]\nProfile: [text]\nOverall: [text]\nNEST: [text]"

    ADF courses:
      - qfi = "Yes" (boolean QFI presence flag) -> use empty string for comment text
      - overall = duplicate of grade number -> use empty string
      - nest = "0" placeholder -> no NEST times exist for any trainee -> empty string

    FIC courses:
      - qfi = actual instructor name -> store as meaningful comment text
      - overall = duplicate of grade number -> use empty string
      - nest = "0" placeholder -> empty string
    """
    # QFI: ADF "Yes" is a boolean flag not a comment; FIC contains instructor name
    if course in ADF_COURSES:
        qfi_text = ""  # "Yes" flag only - no actual comment text
    else:
        # FIC: QFI column holds instructor name - useful contextual info
        qfi_text = str(qfi).strip() if qfi and str(qfi).strip() not in ("", "0", "None") else ""

    # Weather: always real data
    weather_text = str(weather).strip() if weather and str(weather).strip() not in ("", "0", "None") else ""

    # Profile: always real data
    profile_text = str(profile).strip() if profile and str(profile).strip() not in ("", "0", "None") else ""

    # Overall: contains duplicate grade number (e.g., 3) not actual comment text
    overall_text = ""  # Grade number duplicate - not meaningful as comment

    # NEST: always 0 placeholder - no NEST times for any trainee across all courses
    nest_text = ""  # No NEST timing data exists

    # Build structured string
    comment_str = (
        f"QFI: {qfi_text}\n"
        f"Weather: {weather_text}\n"
        f"Profile: {profile_text}\n"
        f"Overall: {overall_text}\n"
        f"NEST: {nest_text}"
    )

    # Return None if all sections are empty (saves storage space)
    if not any([qfi_text, weather_text, profile_text]):
        return None

    return comment_str

def derive_course_context(event_code: str, course_name: str) -> dict:
    """
    Derive syllabus phase and event sequence from the event code.

    Event code patterns:
      ADF courses: "BGF MB1" -> flightNumber="MB1", phase="MB", seq=1
                   "BGF BGF11" -> flightNumber="BGF11", phase="BGF", seq=11
                   "BIF BIF1" -> flightNumber="BIF1", phase="BIF", seq=1
      FIC courses: "FIC GND1" -> flightNumber="GND1", phase="GND", seq=1
                   "FIC FLT1" -> flightNumber="FLT1", phase="FLT", seq=1
    """
    if not event_code:
        return {"flightNumber": "", "syllabusPhase": None, "eventSequence": None}

    s = str(event_code).strip()

    # Extract the meaningful code after the prefix space (e.g., "BGF MB1" -> "MB1")
    parts = s.split()
    if len(parts) >= 2:
        flight_number = parts[-1]  # Last token is the actual event code
    else:
        flight_number = s

    # Determine syllabusPhase from the flight_number prefix
    phase_match = re.match(r'^([A-Z]+)', flight_number)
    phase = phase_match.group(1) if phase_match else None

    # Extract numeric sequence from the flight_number (e.g., "BGF11" -> 11, "MB1" -> 1)
    seq_match = re.search(r'(\d+)$', flight_number)
    sequence = int(seq_match.group(1)) if seq_match else None

    return {
        "flightNumber": flight_number,
        "syllabusPhase": phase,
        "eventSequence": sequence,
    }

def build_element_scores(row, max_cols: int) -> list:
    """
    Build the 22-element scores JSON array matching Pt051Assessment.scores interface.
    [{"element": "Airmanship", "grade": "4", "comment": "..."}, ...]
    """
    scores = []
    for elem in ELEMENTS:
        score_col   = elem["score_col"]
        comment_col = elem["comment_col"]

        # Get score value (1-based col index, row is 0-based tuple)
        raw_grade   = row[score_col - 1]   if score_col   <= max_cols else None
        raw_comment = row[comment_col - 1] if comment_col <= max_cols else None

        grade   = parse_element_grade(raw_grade)
        comment = str(raw_comment).strip() if raw_comment and str(raw_comment).strip() not in ("", "None", "0") else ""

        scores.append({
            "element": elem["name"],
            "grade":   grade,
            "comment": comment,
        })
    return scores

def get_trainee_id(trainee_name: str, course: str) -> str:
    """
    Generate a deterministic traineeId from trainee name + course.
    Uses MD5 hash to ensure same trainee always gets same ID.
    Format: "tp_[8-char-hash]" (tp = trainee performance)
    NOTE: This will be replaced with actual Trainee.id from DB when app links live data.
    """
    key = f"{trainee_name}::{course}".lower().strip()
    h = hashlib.md5(key.encode()).hexdigest()[:12]
    return f"tp_{h}"

# ---------------------------------------------------------------------------
# Core Import Logic
# ---------------------------------------------------------------------------

def process_sheet(ws, sheet_name: str, course: str, errors: list, stats: dict) -> list:
    """
    Process a single trainee sheet and return a list of record dicts.
    Each dict maps directly to a TraineePerformance DB row.
    """
    records = []
    max_cols = ws.max_column

    # Read all rows as values
    all_rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))

    for row_offset, row in enumerate(all_rows):
        actual_row = DATA_START_ROW + row_offset

        # Skip empty rows (Col B must have an event code)
        event_code_raw = row[COL_EVENT_CODE - 1] if len(row) >= COL_EVENT_CODE else None
        if not event_code_raw:
            continue

        event_code = str(event_code_raw).strip()

        # Get trainee name (Col D)
        trainee_name_raw = row[COL_TRAINEE_NAME - 1] if len(row) >= COL_TRAINEE_NAME else None
        trainee_name = str(trainee_name_raw).strip() if trainee_name_raw else sheet_name

        # Derive course context (flightNumber, syllabusPhase, eventSequence)
        ctx = derive_course_context(event_code, course)

        # Parse date
        date_raw = row[COL_DATE - 1] if len(row) >= COL_DATE else None
        parsed_date = parse_date(date_raw, course)
        if not parsed_date:
            errors.append(f"{course}/{sheet_name} row {actual_row}: Invalid date '{date_raw}'")
            stats["date_errors"] += 1

        # Parse instructor
        instructor_raw = row[COL_INSTRUCTOR - 1] if len(row) >= COL_INSTRUCTOR else None
        instructor_name = str(instructor_raw).strip() if instructor_raw else ""

        # Parse grading fields
        overall_grade  = parse_grade(row[COL_OVERALL_GRADE - 1]  if len(row) >= COL_OVERALL_GRADE  else None)
        overall_result = parse_overall_result(row[COL_OVERALL_RESULT - 1] if len(row) >= COL_OVERALL_RESULT else None)
        dco_result_raw = row[COL_DCO_RESULT - 1] if len(row) >= COL_DCO_RESULT else None
        dco_result     = str(dco_result_raw).strip() if dco_result_raw else ""

        # Validate DCO result
        if dco_result not in VALID_DCO_RESULTS:
            errors.append(f"{course}/{sheet_name} row {actual_row}: Unknown dcoResult '{dco_result}' -> set to ''")
            dco_result = ""
            stats["dco_errors"] += 1

        # Parse timing fields
        start_time = parse_time_to_decimal(row[COL_START_TIME - 1] if len(row) >= COL_START_TIME else None)
        duration   = parse_duration(row[COL_DURATION - 1] if len(row) >= COL_DURATION else None, course)
        end_time   = parse_time_to_decimal(row[COL_END_TIME - 1] if len(row) >= COL_END_TIME else None)

        # Parse status flags
        is_completed = parse_bool(row[COL_IS_COMPLETED - 1] if len(row) >= COL_IS_COMPLETED else None)

        # Parse comment fields
        qfi     = row[COL_QFI - 1]            if len(row) >= COL_QFI            else None
        weather = row[COL_WEATHER - 1]         if len(row) >= COL_WEATHER         else None
        profile = row[COL_PROFILE - 1]         if len(row) >= COL_PROFILE         else None
        overall = row[COL_OVERALL_COMMENT - 1] if len(row) >= COL_OVERALL_COMMENT else None
        nest    = row[COL_NEST - 1]            if len(row) >= COL_NEST            else None

        comments = build_comments(qfi, weather, profile, overall, nest, course)

        # Build element scores JSON array
        element_scores = build_element_scores(row, max_cols)

        # Parse ground school assessment
        is_gs      = parse_bool(row[COL_IS_ASSESSMENT - 1] if len(row) >= COL_IS_ASSESSMENT else None)
        gs_result  = parse_ground_school_result(row[COL_GS_RESULT - 1] if len(row) >= COL_GS_RESULT else None)

        # Generate unique IDs
        record_id = generate_cuid()
        event_id  = generate_event_id()
        trainee_id = get_trainee_id(trainee_name, course)

        # Build event description
        event_desc_raw = row[COL_EVENT_DESC - 1] if len(row) >= COL_EVENT_DESC else None
        event_desc = str(event_desc_raw).strip() if event_desc_raw else None

        # Build the record
        record = {
            "id":                       record_id,
            "traineeId":                trainee_id,
            "traineeFullName":          trainee_name,
            "eventId":                  event_id,
            "eventCode":                event_code,
            "flightNumber":             ctx["flightNumber"],
            "eventDescription":         event_desc,
            "date":                     parsed_date or "1900-01-01",  # Fallback for bad dates
            "instructorName":           instructor_name,
            "instructorId":             None,
            "overallGrade":             overall_grade,
            "overallResult":            overall_result,
            "dcoResult":                dco_result,
            "startTime":                start_time,
            "duration":                 duration,
            "endTime":                  end_time,
            "comments":                 comments,
            "elementScores":            element_scores,
            "isCompleted":              is_completed,
            "isGroundSchoolAssessment": is_gs,
            "groundSchoolResult":       gs_result,
            "course":                   course,
            "syllabusPhase":            ctx["syllabusPhase"],
            "eventSequence":            ctx["eventSequence"],
            "createdBy":                "import_script",
            "updatedBy":                "import_script",
        }

        records.append(record)
        stats["records"] += 1

        # Small sleep to ensure unique eventIds (timestamp-based)
        time.sleep(0.001)

    return records

def process_spreadsheet(filepath: str, course: str) -> tuple[list, list]:
    """
    Process all sheets in a spreadsheet.
    Returns (all_records, errors).
    """
    print(f"\n{'='*60}")
    print(f"Processing: {course} ({filepath.split('/')[-1]})")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_records = []
    errors = []
    stats = {"records": 0, "date_errors": 0, "dco_errors": 0}

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        print(f"  [{sheet_idx+1:2d}/{len(wb.sheetnames)}] {sheet_name}...", end=" ")

        records = process_sheet(ws, sheet_name, course, errors, stats)
        all_records.extend(records)
        print(f"{len(records)} records")

    wb.close()

    print(f"\n  ✅ {course}: {stats['records']} records from {len(wb.sheetnames)} trainees")
    if stats["date_errors"]:
        print(f"  ⚠️  Date errors: {stats['date_errors']}")
    if stats["dco_errors"]:
        print(f"  ⚠️  DCO errors: {stats['dco_errors']}")
    if errors:
        print(f"  ⚠️  Total errors: {len(errors)}")

    return all_records, errors

def save_records_to_json(all_records: list, output_path: str):
    """
    Save all records to a JSON file for inspection and/or DB insertion.
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(all_records, f, indent=2, default=str, ensure_ascii=False)
    print(f"\n💾 Saved {len(all_records)} records to: {output_path}")

def validate_records(all_records: list) -> list:
    """
    Validate all records and return a list of validation errors.
    """
    validation_errors = []
    event_ids_seen = set()

    for i, r in enumerate(all_records):
        # Check required fields
        for field in ["id", "traineeId", "traineeFullName", "eventId",
                      "eventCode", "flightNumber", "date", "instructorName",
                      "overallGrade", "elementScores"]:
            if not r.get(field):
                validation_errors.append(f"Record {i}: Missing required field '{field}'")

        # Check eventId uniqueness
        eid = r.get("eventId")
        if eid in event_ids_seen:
            validation_errors.append(f"Record {i}: Duplicate eventId '{eid}'")
        event_ids_seen.add(eid)

        # Check overall grade validity
        if r.get("overallGrade") not in VALID_OVERALL_GRADES:
            validation_errors.append(f"Record {i}: Invalid overallGrade '{r.get('overallGrade')}'")

        # Check element count
        if len(r.get("elementScores", [])) != 22:
            validation_errors.append(
                f"Record {i}: Expected 22 elements, got {len(r.get('elementScores', []))}"
            )

        # Check date format
        date_val = r.get("date", "")
        if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_val):
            validation_errors.append(f"Record {i}: Invalid date format '{date_val}'")

    return validation_errors

def generate_sql_inserts(records: list, output_path: str):
    """
    Generate SQL INSERT statements for all records.
    Outputs a .sql file that can be run directly against PostgreSQL.
    """
    def sql_str(v):
        if v is None:
            return "NULL"
        s = str(v).replace("'", "''")
        return f"'{s}'"

    def sql_bool(v):
        return "TRUE" if v else "FALSE"

    def sql_num(v):
        if v is None:
            return "NULL"
        return str(v)

    def sql_json(v):
        if v is None:
            return "NULL"
        j = json.dumps(v, ensure_ascii=False).replace("'", "''")
        return f"'{j}'::jsonb"

    lines = [
        "-- TraineePerformance Data Import",
        f"-- Generated: {datetime.now().isoformat()}",
        f"-- Total records: {len(records)}",
        "-- Run this file against your PostgreSQL database",
        "",
        "BEGIN;",
        "",
    ]

    for r in records:
        line = (
            f"INSERT INTO \"TraineePerformance\" ("
            f"\"id\", \"traineeId\", \"traineeFullName\", \"eventId\", "
            f"\"eventCode\", \"flightNumber\", \"eventDescription\", \"date\", "
            f"\"instructorName\", \"instructorId\", "
            f"\"overallGrade\", \"overallResult\", \"dcoResult\", "
            f"\"startTime\", \"duration\", \"endTime\", "
            f"\"comments\", \"elementScores\", "
            f"\"isCompleted\", \"isGroundSchoolAssessment\", \"groundSchoolResult\", "
            f"\"course\", \"syllabusPhase\", \"eventSequence\", "
            f"\"createdBy\", \"updatedBy\", \"updatedAt\""
            f") VALUES ("
            f"{sql_str(r['id'])}, "
            f"{sql_str(r['traineeId'])}, "
            f"{sql_str(r['traineeFullName'])}, "
            f"{sql_str(r['eventId'])}, "
            f"{sql_str(r['eventCode'])}, "
            f"{sql_str(r['flightNumber'])}, "
            f"{sql_str(r['eventDescription'])}, "
            f"{sql_str(r['date'])}, "
            f"{sql_str(r['instructorName'])}, "
            f"{sql_str(r['instructorId'])}, "
            f"{sql_str(r['overallGrade'])}, "
            f"{sql_str(r['overallResult'])}, "
            f"{sql_str(r['dcoResult'])}, "
            f"{sql_num(r['startTime'])}, "
            f"{sql_num(r['duration'])}, "
            f"{sql_num(r['endTime'])}, "
            f"{sql_str(r['comments'])}, "
            f"{sql_json(r['elementScores'])}, "
            f"{sql_bool(r['isCompleted'])}, "
            f"{sql_bool(r['isGroundSchoolAssessment'])}, "
            f"{sql_num(r['groundSchoolResult'])}, "
            f"{sql_str(r['course'])}, "
            f"{sql_str(r['syllabusPhase'])}, "
            f"{sql_num(r['eventSequence'])}, "
            f"{sql_str(r['createdBy'])}, "
            f"{sql_str(r['updatedBy'])}, "
            f"NOW()"
            f") ON CONFLICT (\"eventId\") DO NOTHING;"
        )
        lines.append(line)

    lines.extend([
        "",
        "COMMIT;",
        "",
        f"-- Verify import:",
        f"-- SELECT course, COUNT(*) as records FROM \"TraineePerformance\" GROUP BY course ORDER BY course;",
    ])

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f"💾 SQL INSERT file saved to: {output_path}")

def print_summary(all_records: list):
    """Print a comprehensive import summary."""
    from collections import Counter

    print(f"\n{'='*60}")
    print("IMPORT SUMMARY")
    print(f"{'='*60}")
    print(f"Total records: {len(all_records)}")

    # By course
    course_counts = Counter(r['course'] for r in all_records)
    print(f"\nBy course:")
    for course, count in sorted(course_counts.items()):
        trainees = len(set(r['traineeFullName'] for r in all_records if r['course'] == course))
        events   = count // trainees if trainees else 0
        print(f"  {course:<10} {count:>5} records  ({trainees} trainees × ~{events} events)")

    # By syllabus phase
    phase_counts = Counter(r['syllabusPhase'] for r in all_records)
    print(f"\nBy syllabus phase:")
    for phase, count in sorted(phase_counts.items()):
        print(f"  {str(phase):<12} {count:>5} records")

    # Grade distribution
    grade_counts = Counter(r['overallGrade'] for r in all_records)
    print(f"\nOverall grade distribution:")
    for grade in sorted(grade_counts.keys()):
        print(f"  Grade {grade:<8} {grade_counts[grade]:>5} records")

    # DCO distribution
    dco_counts = Counter(r['dcoResult'] for r in all_records)
    print(f"\nDCO result distribution:")
    for dco, count in dco_counts.items():
        print(f"  {str(dco):<8} {count:>5} records")

    # Completion status
    completed = sum(1 for r in all_records if r['isCompleted'])
    print(f"\nCompletion status:")
    print(f"  Completed:   {completed:>5}")
    print(f"  Incomplete:  {len(all_records) - completed:>5}")

    print(f"\n{'='*60}")

# ---------------------------------------------------------------------------
# Main Entry Point
# ---------------------------------------------------------------------------

def main():
    print("TraineePerformance Import Script")
    print("="*60)
    print(f"Processing {len(SPREADSHEETS)} spreadsheets...")

    all_records = []
    all_errors  = []

    for filepath, course in SPREADSHEETS:
        if not os.path.exists(filepath):
            print(f"⚠️  File not found: {filepath}")
            continue
        records, errors = process_spreadsheet(filepath, course)
        all_records.extend(records)
        all_errors.extend(errors)

    # Validate all records
    print(f"\n{'='*60}")
    print("Validating records...")
    validation_errors = validate_records(all_records)
    if validation_errors:
        print(f"⚠️  {len(validation_errors)} validation errors found:")
        for e in validation_errors[:20]:
            print(f"  - {e}")
        if len(validation_errors) > 20:
            print(f"  ... and {len(validation_errors) - 20} more")
    else:
        print(f"✅ All {len(all_records)} records validated successfully")

    # Print summary
    print_summary(all_records)

    # Save outputs
    output_dir = "/workspace/DFP-NEO-V2/migration-scripts"

    # JSON output (for inspection and Prisma-based insertion)
    json_path = f"{output_dir}/trainee_performance_import.json"
    save_records_to_json(all_records, json_path)

    # SQL output (for direct PostgreSQL execution)
    sql_path = f"{output_dir}/trainee_performance_data.sql"
    generate_sql_inserts(all_records, sql_path)

    # Error log
    if all_errors:
        error_path = f"{output_dir}/import_errors.log"
        with open(error_path, 'w') as f:
            f.write(f"Import Errors - {datetime.now().isoformat()}\n")
            f.write("="*60 + "\n")
            for e in all_errors:
                f.write(f"{e}\n")
        print(f"⚠️  Errors logged to: {error_path}")

    print(f"\n✅ Import preparation complete!")
    print(f"   Records ready: {len(all_records)}")
    print(f"   JSON file:     {json_path}")
    print(f"   SQL file:      {sql_path}")
    print(f"\nNext step: Apply migration SQL then run the data SQL against your database.")

if __name__ == "__main__":
    main()